import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import linregress
from scipy.io.wavfile import read
from scipy.signal.windows import hann
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

def EnergyCalc(y,sf,ef):
    N=len(y)
    fft=np.fft.fft(y)
    fft_abs=2*(np.abs(fft)/N)
    correction_factor = np.sum(window) / len(window)
    corrected_signal = fft_abs / correction_factor
    fft_log=20*np.log10(corrected_signal)
    dt=1/sr
    freq = np.fft.fftfreq(N, d=dt)
    df=sr/N
    start=int(sf/df)
    end=int(ef/df)
    data=fft_abs[start:end]
    data=data[data>-96]
    energy=sum(data)
    print(f"energy:{energy:.3f}")
    # plt.figure(figsize=(8, 6))
    # plt.plot(freq[0:int(N/2)],fft_log[0:int(N/2)])
    # plt.xlabel("Frequency(Hz)")
    # plt.ylabel("Amplitude")
    # plt.xscale("log")
    # plt.ylim(-100,10)
    # plt.grid()
    # plt.show()
    return energy

second=9.8
n_fft=2048
window=hann(n_fft)
drive=[0,5,10]
tone=[0,5,10]
level=[0,5,10]
dir_path=input("分析するディレクトリのパスを入力してください。\n")
save_file_name=input("保存するファイル名を入力してください。\n")
f0=int(input("基本周波数を入力してください"))
energy_list=[]
fname_list=[]
# sf=2*f0
sf=4000
ef=22050

for d in drive:
    for t in tone:
        for l in level:
            if l != 0 or d == t == l ==0: 
                filename=f"[{d}, {t}, {l}].wav"
                fname_list.append(filename)
                sr,y=read(f"{dir_path}\{filename}")
                y=y/32768
                signal=y[int(second*sr):int((second*sr)+n_fft)]
                w_signal=window*signal
                y=y[9*sr:int(10*sr)]
                energy_list.append(EnergyCalc(w_signal,sf,ef))
df=pd.DataFrame(energy_list,columns=["エネルギー"],index=fname_list)
df.to_excel(f"energy/{save_file_name}.xlsx")
wb = load_workbook(f"energy/{save_file_name}.xlsx")
ws = wb.active
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if isinstance(cell.value, (int, float)):
                max_length = max(max_length, len(f"{cell.value:.2f}") + 4)  
            elif cell.value:  
                max_length = max(max_length, len(str(cell.value))-2)
        except:
            pass
    adjusted_width = max_length + 3 
    ws.column_dimensions[column].width = adjusted_width
for cell in ws['A']:  
    cell.alignment = Alignment(horizontal='left')
wb.save(f"energy/{save_file_name}.xlsx") 