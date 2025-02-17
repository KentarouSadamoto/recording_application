import numpy as np
import matplotlib.pyplot as plt
from scipy.io.wavfile import read
from scipy.signal.windows import hann
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import make_dir as md

def compute_spectral_slope(signal, fs, f_min=100, f_max=None):
    if f_max is None:
        f_max = fs / 2  # ナイキスト周波数
    # FFTを計算
    N=len(signal)
    freqs = np.fft.rfftfreq(N, d=1/fs)
    fft_abs = np.abs(np.fft.rfft(signal))/N
    correction_factor = np.sum(window) / len(window)
    corrected_signal = fft_abs / correction_factor
    # dBスケールへ変換
    spectrum_db = 20 * np.log10(corrected_signal) 

    # 指定範囲の周波数を抽出
    # f_min=int(f_min)
    valid_idx = (freqs >= f_min) & (freqs <= f_max)
    freqs = freqs[valid_idx]
    spectrum_db = spectrum_db[valid_idx]
    # 周波数をオクターブスケールに変換
    log_freqs = np.log2(freqs)
    #-96dB以下のデータを削除
    spectrum_db_96=spectrum_db[spectrum_db > -96]
    log_freqs_96=log_freqs[spectrum_db > -96]
    # 最小二乗法で直線回帰
    slope, intercept = np.polyfit(log_freqs_96, spectrum_db_96, 1)

    # グラフ描画
    plt.figure(figsize=(8, 5))
    plt.plot(freqs, spectrum_db, label="spectrum", color="blue", alpha=0.7)
    plt.plot(freqs, slope * np.log2(freqs) + intercept, label=f"spectral slope ({slope:.2f} dB/oct)", color="red", linestyle="--")
    plt.xscale("log")  # 周波数軸を対数スケールに
    plt.xlabel("Frequency (Hz)")
    plt.ylabel("Amplitude (dB)")
    plt.legend()
    plt.ylim(-100,0)
    plt.grid(True, which="both", linestyle="--", alpha=0.5)
    plt.savefig(f"octave_spectral_slope\{save_file_name}\DRIVE{d}TONE{t}LEVEL{l}.png")

    return slope  # dB/oct の値

# サンプル信号を生成（ホワイトノイズ）
n_fft=2048
drive=[0,5,10]
tone=[0,5,10]
level=[0,5,10]
dir_path=input("分析するディレクトリのパスを入力してください。\n")
save_file_name=input("保存するファイル名を入力してください。\n")
second=int(input("何秒時点のデータを分析しますか"))
f_min,f_max=int(input("分析に使う最低周波数は何Hzですか。")),int(input("分析に使う最高周波数は何Hzですか。"))
slope_list=[]
fname_list=[]
md.make_dir(f"octave_spectral_slope",save_file_name)

for d in drive:
    for t in tone:
        for l in level:
            if l != 0 or d == t == l ==0: 
                filename=f"[{d}, {t}, {l}].wav"
                sr,y=read(f"{dir_path}\{filename}")
                fname_list.append(f"DRIVE={d},TONE={t},LEVEL={l}.wav")
                y=y/32768
                signal=y[int(second*sr):int((second*sr)+n_fft)]
                window=hann(n_fft)
                w_signal=window*signal
                slope = compute_spectral_slope(w_signal, sr,f_min=f_min,f_max=f_max)
                slope_list.append(slope)

df=pd.DataFrame(slope_list,columns=["スペクトル傾斜"],index=fname_list)
df.to_excel(f"octave_spectral_slope/{save_file_name}.xlsx")
wb = load_workbook(f"octave_spectral_slope/{save_file_name}.xlsx")
ws = wb.active
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if isinstance(cell.value, (int, float)):
                max_length = max(max_length, len(f"{cell.value:.2f}") + 4)  
            elif cell.value:  
                max_length = max(max_length, len(str(cell.value))-2)
        except:
            pass
    adjusted_width = max_length + 3 
    ws.column_dimensions[column].width = adjusted_width
for cell in ws['A']:  
    cell.alignment = Alignment(horizontal='left')
wb.save(f"octave_spectral_slope/{save_file_name}.xlsx")