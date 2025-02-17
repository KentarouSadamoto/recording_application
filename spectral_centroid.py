import librosa
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

file_path=input("分析するディレクトリのパスを入力してください。\n")
file_name=input("保存するファイル名を入力してください。\n")
drive=[0,5,10]
tone=[0,5,10]
level=[0,5,10]
index=[]
cent_list=[]
duration=44100

for i in drive:
    for j in tone:
        for k in level:
            if k != 0 or i == j == k ==0: #エフェクターをONにした時の各パラメータおよびバイパス時[0, 0, 0](gain:0,tone:0,level:0)のwavファイルを分析する。
                filename=f"[{i}, {j}, {k}].wav"
                index.append(filename)
                y, sr = librosa.load(fr"{file_path}\{filename}",sr=44100)
                nine_sec=y[9*sr:9*sr+duration]
                cent_9sec = librosa.feature.spectral_centroid(y=nine_sec, sr=sr)
                cent_list.append(np.round(np.average(cent_9sec[0,][3:-2])))

df=pd.DataFrame(cent_list,columns=["スペクトル重心"],index=index)
df.to_excel(f"spectral_centroid/{file_name}.xlsx")
wb = load_workbook(f"spectral_centroid/{file_name}.xlsx")
ws = wb.active
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if isinstance(cell.value, (int, float)):
                max_length = max(max_length, len(f"{cell.value:.2f}") + 4)  
            elif cell.value:  
                max_length = max(max_length, len(str(cell.value))-2)
        except:
            pass
    adjusted_width = max_length + 3 
    ws.column_dimensions[column].width = adjusted_width
for cell in ws['A']:  
    cell.alignment = Alignment(horizontal='left')
wb.save(f"spectral_centroid/{file_name}.xlsx")

