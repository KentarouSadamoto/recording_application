import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import linregress
from scipy.io.wavfile import read
from scipy.signal.windows import hann
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import make_dir as md

def linear_regression(y, f0):
    # FFT
    N = len(y)
    fft = np.fft.fft(y)
    fft_abs = 2 * (np.abs(fft) / N)
    correction_factor = np.sum(window) / len(window)
    corrected_signal = fft_abs / correction_factor
    fft_log = 20 * np.log10(corrected_signal)
    dt = 1 / sr
    freq = np.fft.fftfreq(N, d=dt)

    # 基本音と倍音の振幅を取り出す
    df = sr / N
    idx_array = np.array([])
    for i in range(f0, 22050, f0):
        idx = int(i / df)
        idx_array = np.append(idx_array, idx)
    idx_array = idx_array.astype(int)
    amplitude = fft_log[idx_array]

    # 負の値やノイズを削除
    deletelist = []
    for i in range(len(amplitude)):
        if amplitude[i] < -96:  # -96dB以下を削除
            deletelist.append(i)
    idx_array = np.delete(idx_array, deletelist)
    amplitude = np.delete(amplitude, deletelist)

    # 周波数と振幅を対数変換
    freq_log = np.log10(idx_array * df)
    print(freq_log)
    slope, intercept, _, _, _ = linregress(freq_log, amplitude)

    # 直線（回帰直線）の計算
    y_pred = slope * freq_log + intercept

    # プロット
    plt.figure(figsize=(8, 7))
    plt.plot(freq[0:int(N / 2)], fft_log[0:int(N / 2)], label="Data", color="blue")
    plt.scatter(idx_array * df, amplitude, color="green", label="sample")
    plt.plot(idx_array * df, y_pred, label=f"Regression: y = {slope:.2f} * log10(f) + {intercept:.2f}", color="red")

    # 軸の設定（横軸は対数スケール）
    plt.xscale('log')
    plt.xlabel("Frequency (Hz)")
    plt.ylabel("Amplitude (dB)")
    plt.ylim(-100, 10)
    plt.legend()
    plt.grid(which='both', linestyle='--', linewidth=0.5)
    plt.savefig(f"spectral_slope\{save_file_name}\DRIVE{d}TONE{t}LEVEL{l}.png")
    plt.close()

    return slope


second=9.8
n_fft=2048
drive=[0,5,10]
tone=[0,5,10]
level=[0,5,10]
dir_path=input("分析するディレクトリのパスを入力してください。\n")
save_file_name=input("保存するファイル名を入力してください。\n")
f0=int(input("基本周波数を入力してください"))
slope_list=[]
fname_list=[]
md.make_dir(f"spectral_slope",save_file_name)

for d in drive:
    for t in tone:
        for l in level:
            if l != 0 or d == t == l ==0: 
                filename=f"[{d}, {t}, {l}].wav"
                sr,y=read(f"{dir_path}\{filename}")
                fname_list.append(f"DRIVE={d},TONE={t},LEVEL={l}")
                y=y/32768
                signal=y[int(second*sr):int((second*sr)+n_fft)]
                window=hann(n_fft)
                w_signal=window*signal
                slope=linear_regression(w_signal,f0)
                slope_list.append(slope)

df=pd.DataFrame(slope_list,columns=["スペクトル傾斜"],index=fname_list)
df.to_excel(f"spectral_slope/{save_file_name}.xlsx")
wb = load_workbook(f"spectral_slope/{save_file_name}.xlsx")
ws = wb.active
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if isinstance(cell.value, (int, float)):
                max_length = max(max_length, len(f"{cell.value:.2f}") + 4)  
            elif cell.value:  
                max_length = max(max_length, len(str(cell.value))-2)
        except:
            pass
    adjusted_width = max_length + 3 
    ws.column_dimensions[column].width = adjusted_width
for cell in ws['A']:  
    cell.alignment = Alignment(horizontal='left')
wb.save(f"spectral_slope/{save_file_name}.xlsx")