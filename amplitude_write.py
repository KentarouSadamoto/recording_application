import numpy as np
import pandas as pd
from scipy.io.wavfile import read
from openpyxl import load_workbook
from openpyxl.styles import Alignment

sr=44100
def amplitude_analysis(data):
        N=len(data)
        data=data/np.iinfo(np.int16).max
        fft=np.fft.fft(data)
        fft_abs=2*(np.abs(fft)/N)
        fft_log=20*np.log(fft_abs/1)
        index.append(f"[{i}, {j}, {k}].wav")
        fft_values = [round(fft_log[f_idx],2) for f_idx in freq_idx]
        calc_time=len(fft_values)
        for idx in range(1,calc_time):
            fft_values.append(round(fft_values[0]-fft_values[idx],2))
        return fft_values

file_path=input("分析するディレクトリのパスを入力してください。\n")
file_name=input("保存するファイル名を入力してください。\n")
duration=1*sr
f0=1000
drive=[0,5,10]
tone=[0,5,10]
level=[0,5,10]
index=[]
result=[]
freq_idx = [int(f * duration / sr) for f in [f0, 2*f0, 3*f0, 4*f0, 5*f0, 6*f0, 7*f0,8*f0]]
for i in drive:
    for j in tone:
        for k in level:
            if k != 0 or i == j == k ==0: #エフェクターをONにした時の各パラメータおよびバイパス時[0, 0, 0](gain:0,tone:0,level:0)のwavファイルを分析する。
                filename=f"[{i}, {j}, {k}].wav"
                sr, data = read(fr"{file_path}\{filename}")
                sec9=data[int(9*sr):int(9*sr+duration)]
                result.append(amplitude_analysis(sec9))
                
df=pd.DataFrame(result,columns=["f0","f2","f3","f4","f5","f6","f7","f8","f0-f2","f0-f3","f0-f4","f0-f5","f0-f6","f0-f7","f0-f8"],index=index)
df.to_excel(f"amplitude_values/{file_name}.xlsx",sheet_name="test")

wb = load_workbook(f"amplitude_values/{file_name}.xlsx")
ws = wb.active
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            # 数値の場合、少し余裕をもって幅を設定
            if isinstance(cell.value, (int, float)):
                max_length = max(max_length, len(f"{cell.value:.2f}") + 4)  # 小数点以下2桁で計算し、余裕を追加
            elif cell.value:  
                max_length = max(max_length, len(str(cell.value)) - 2)
        except:
            pass
    adjusted_width = max_length + 3 
    ws.column_dimensions[column].width = adjusted_width
for cell in ws['A']:  
    cell.alignment = Alignment(horizontal='left')
wb.save(f"amplitude_values/{file_name}.xlsx")