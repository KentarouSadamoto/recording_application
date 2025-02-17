import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import linregress
from scipy.io.wavfile import read
from scipy.signal.windows import hann
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import make_dir as md

def linear_regression(y,f0):
    #FFT
    N=len(y)
    fft=np.fft.fft(y)
    fft_abs=2*(np.abs(fft)/N)
    correction_factor = np.sum(window) / len(window)
    corrected_signal = fft_abs / correction_factor
    fft_log=20*np.log10(corrected_signal)
    dt=1/sr
    freq = np.fft.fftfreq(N, d=dt)
    #基本音、倍音の振幅を取り出す。例：F0=440Hzの場合、440Hz,880Hz,...の振幅が入っているインデックスを計算して取り出す。
    df=sr/N
    idx_array=np.array([])
    
    # for i in range(f0,22050,f0):
    # for i in range(f0,22050,2*f0):#偶数倍音飛ばすver
    calc=f0
    cnt=0
    while calc <= 22050:
        idx=int(calc/df)#FFTのデータは、周波数分解能dfの整数倍の周波数の時の振幅データが入っており、取り出したい周波数に最も近い周波数を計算するため、(取り出したい周波数/df)をして小数点を切り捨てる計算をしている。
        idx_array=np.append(idx_array,idx,)
        print("calc",calc)
        cnt+=1
        calc=calc*10

    idx_array=(idx_array).astype(int)
    amplitude=fft_log[idx_array]
    deletelist=[]
    for i in range(len(amplitude)):
        if amplitude[i] < -96:#int16bitの分解能は-96dBあたりまでのため-96dB以下のデータを計算に含まないように削除しています。
            deletelist.append(i)
    idx_array=np.delete(idx_array,deletelist)
    amplitude=np.delete(amplitude,deletelist)
    x=idx_array*df
    print("x",x)
    print("amplitude",amplitude)
    slope, intercept, _,_,_= linregress(x, amplitude)
    # print(f"傾き (slope): {slope:.5f}")
    # print(f"切片 (intercept): {intercept:.5f}")
    y_pred = slope * x + intercept
    plt.figure(figsize=(8, 6))
    plt.plot(freq[0:int(N/2)],fft_log[0:int(N/2)],label="data", color="blue")
    plt.scatter(x, amplitude,color="green")
    plt.plot(x, y_pred, label=f"linear regression: y = {slope}x + {intercept}", color="red")
    plt.xlabel("Frequency(Hz)")
    plt.ylabel("Amplitude(dB)")
    plt.ylim(-130,10)
    plt.legend()
    plt.xscale("log")
    plt.grid()
    plt.savefig(f"spectral_slope\{save_file_name}\DRIVE{d}TONE{t}LEVEL{l}.png")
    plt.close()
    return slope

second=9.8
n_fft=2048
drive=[0,5,10]
tone=[0,5,10]
level=[0,5,10]
dir_path=input("分析するディレクトリのパスを入力してください。\n")
save_file_name=input("保存するファイル名を入力してください。\n")
f0=int(input("基本周波数を入力してください"))
slope_list=[]
fname_list=[]
md.make_dir(f"spectral_slope",save_file_name)

for d in drive:
    for t in tone:
        for l in level:
            if l != 0 or d == t == l ==0: 
                filename=f"[{d}, {t}, {l}].wav"
                sr,y=read(f"{dir_path}\{filename}")
                fname_list.append(f"DRIVE={d},TONE={t},LEVEL={l}")
                y=y/32768
                signal=y[int(second*sr):int((second*sr)+n_fft)]
                window=hann(n_fft)
                w_signal=window*signal
                slope=linear_regression(w_signal,f0)
                slope_list.append(slope)

df=pd.DataFrame(slope_list,columns=["スペクトル傾斜"],index=fname_list)
df.to_excel(f"spectral_slope/{save_file_name}.xlsx")
wb = load_workbook(f"spectral_slope/{save_file_name}.xlsx")
ws = wb.active
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if isinstance(cell.value, (int, float)):
                max_length = max(max_length, len(f"{cell.value:.2f}") + 4)  
            elif cell.value:  
                max_length = max(max_length, len(str(cell.value))-2)
        except:
            pass
    adjusted_width = max_length + 3 
    ws.column_dimensions[column].width = adjusted_width
for cell in ws['A']:  
    cell.alignment = Alignment(horizontal='left')
wb.save(f"spectral_slope/{save_file_name}.xlsx")